"""Step 7: Generation QA report workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    PipelineSummary,
    RaciCandidate,
    RawScopeSignal,
    NormalizedSignal,
    SelectedDocument,
    UncertainMapping,
)
from .utils import safe_excel_value

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, headers: List[str], rows: List[List[Any]]) -> None:
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=safe_excel_value(val))
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def write_qa_report(
    output_path: Path,
    raw_signals: List[RawScopeSignal],
    normalized: List[NormalizedSignal],
    uncertain: List[UncertainMapping],
    candidates: List[RaciCandidate],
    selected: List[SelectedDocument],
    summary: PipelineSummary,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # Scope_raw
    ws = wb.create_sheet("Scope_raw")
    _write_sheet(
        ws,
        [
            "source_pdf",
            "extraction_method",
            "scope_section",
            "discipline_code",
            "chapter_name",
            "confidence",
            "source_pages",
            "evidence_quote",
            "notes",
        ],
        [
            [
                s.source_pdf,
                s.extraction_method,
                s.scope_section,
                s.discipline_code,
                s.chapter_name or "",
                s.confidence,
                ",".join(str(p) for p in s.source_pages),
                s.evidence_quote,
                s.notes,
            ]
            for s in raw_signals
        ],
    )

    # Scope_normalized
    ws = wb.create_sheet("Scope_normalized")
    _write_sheet(
        ws,
        [
            "discipline_code",
            "chapter_name",
            "scope_section",
            "confidence",
            "normalization_method",
            "use_chapter_filter",
            "source_pdf",
            "notes",
        ],
        [
            [
                n.discipline_code,
                n.chapter_name or "",
                n.scope_section,
                n.confidence,
                n.normalization_method,
                n.use_chapter_filter,
                n.source_pdf,
                n.notes,
            ]
            for n in normalized
        ],
    )

    # Candidates
    ws = wb.create_sheet("Candidates")
    _write_sheet(
        ws,
        [
            "rank",
            "title_key",
            "title",
            "discipline_code",
            "chapter_name",
            "type_code",
            "historical_count",
            "avg_confidence",
            "judge_hits",
            "recovery_hits",
        ],
        [
            [
                c.rank,
                c.title_key,
                c.title,
                c.discipline_code,
                c.chapter_name,
                c.type_code,
                c.historical_count,
                c.avg_confidence,
                c.judge_hits,
                c.recovery_hits,
            ]
            for c in candidates
        ],
    )

    # Selected
    ws = wb.create_sheet("Selected")
    _write_sheet(
        ws,
        [
            "title_key",
            "title",
            "discipline_code",
            "chapter_name",
            "type_code",
            "historical_count",
            "avg_confidence",
            "bucket",
            "selection_reason",
        ],
        [
            [
                s.title_key,
                s.title,
                s.discipline_code,
                s.chapter_name,
                s.type_code,
                s.historical_count,
                s.avg_confidence,
                s.bucket,
                s.selection_reason,
            ]
            for s in selected
        ],
    )

    # Without_history
    ws = wb.create_sheet("Without_history")
    without = [s for s in selected if s.bucket == "without_history"]
    _write_sheet(
        ws,
        ["title_key", "title", "discipline_code", "chapter_name", "type_code"],
        [
            [s.title_key, s.title, s.discipline_code, s.chapter_name, s.type_code]
            for s in without
        ],
    )

    # Excluded
    ws = wb.create_sheet("Excluded")
    _write_sheet(
        ws,
        ["raw_discipline", "raw_chapter", "reason", "scope_section", "source_pdf"],
        [
            [u.raw_discipline, u.raw_chapter, u.reason, u.scope_section, u.source_pdf]
            for u in uncertain
        ],
    )

    # Summary
    ws = wb.create_sheet("Summary")
    summary_rows = [
        ["project_name", summary.project_name],
        ["scope_pdfs", "; ".join(summary.scope_pdfs)],
        ["disciplines_found", ", ".join(summary.disciplines_found)],
        ["chapters_found", ", ".join(summary.chapters_found)],
        ["raw_signal_count", summary.raw_signal_count],
        ["normalized_signal_count", summary.normalized_signal_count],
        ["candidate_count", summary.candidate_count],
        ["selected_count", summary.selected_count],
        ["with_history_count", summary.with_history_count],
        ["without_history_count", summary.without_history_count],
        ["duplicates_removed", summary.duplicates_removed],
        ["uncertain_mapping_count", summary.uncertain_mapping_count],
    ]
    _write_sheet(ws, ["metric", "value"], summary_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
