"""Step 12: Build Master Document Register Excel (clean workbook, no template copy)."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn
from openpyxl.worksheet.worksheet import Worksheet

from .models import MdrLineItem
from .utils import safe_excel_value

DATA_START_ROW = 14
HEADER_LAST_ROW = 13
COL_ITEM = 1
COL_DESCRIPTION = 2  # B — unica colonna dati per intestazione unita B13:F13
COL_FORMULA = 7
COL_DISC = 9
COL_TYPE = 10
COL_PROG = 11  # K — progressive per coppia I+J
COL_CATEGORY = 15
COL_WBS = 21  # U
COL_MANHOURS = 24  # X — ore/uomo = giorni timeline × 8
COL_WORKFLOW = 27  # AA
COL_PLANNED_FIRST_ISSUE = 29  # AC — PLANNED FIRST ISSUE (AD non compilata per ora)
# Schedule debug columns (only when schedule.debug_columns=true) — after AI (col 35)
COL_DBG_FIRST = 36  # AJ
SCHEDULE_DEBUG_META_ROW = 12  # project start once, above DBG column headers
SCHEDULE_DEBUG_HEADERS = [
    "DBG_TitleKey",
    "DBG_PredFinishes",
    "DBG_DrivingPred",
    "DBG_PlannedStart",
    "DBG_DurationDays",
    "DBG_PlannedFinish",
    "DBG_MissingPreds",
    "DBG_Flags",
]
# Offsets within SCHEDULE_DEBUG_HEADERS for date-formatted cells.
_DBG_COL_PLANNED_START = 3
_DBG_COL_PLANNED_FINISH = 5
FILTER_HEADER_ROW = 13
FILTER_FIRST_COL = "A"
FILTER_LAST_COL = "AI"  # col 35 — KBS INTERNAL CODE
# Template: B13:F13 merged — filtro descrizione solo su colonna B (offset 1 da A).
FILTER_MERGED_DESC_FIRST_COL = COL_DESCRIPTION
FILTER_MERGED_DESC_LAST_COL = 6  # F
SHEET_NAME = "MDR"
DATE_NUMBER_FORMAT = "mm-dd-yy"


def _copy_cell_style(src, dst) -> None:
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def _copy_template_header(source: Worksheet, target: Worksheet, max_row: int = HEADER_LAST_ROW) -> None:
    """Copy rows 1..max_row (values, styles, merges, dimensions, images) — no print metadata."""
    max_col = max(source.max_column, COL_WORKFLOW)

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        dim = source.column_dimensions.get(letter)
        if dim is not None and dim.width is not None:
            target.column_dimensions[letter].width = dim.width

    for row_idx in range(1, max_row + 1):
        src_rd = source.row_dimensions.get(row_idx)
        if src_rd is not None and src_rd.height is not None:
            target.row_dimensions[row_idx].height = src_rd.height

    for merge in source.merged_cells.ranges:
        if merge.min_row <= max_row and merge.max_row <= max_row:
            target.merge_cells(str(merge))

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            src = source.cell(row=row_idx, column=col_idx)
            dst = target.cell(row=row_idx, column=col_idx, value=src.value)
            _copy_cell_style(src, dst)

    for image in getattr(source, "_images", ()):
        target.add_image(copy(image), image.anchor)


def _reneco_code_formula(row: int, project_prefix: str) -> str:
    """H (originator) is left empty for manual entry; K is the I+J pair progressive."""
    prefix = (project_prefix or "0000").replace('"', '""')
    return f'="{prefix}-"&H{row}&"-"&I{row}&J{row}&"-"&K{row}'


def _line_title(doc: MdrLineItem) -> str:
    return doc.mdr_document_title


def _line_discipline(doc: MdrLineItem) -> str:
    return doc.discipline_code


def _line_type(doc: MdrLineItem) -> str:
    return doc.type_code or ""


def _line_category(doc: MdrLineItem) -> str:
    return doc.category_code or ""


def _line_wbs(doc: MdrLineItem) -> str:
    return doc.discipline_wbs or ""


def _line_workflow(doc: MdrLineItem) -> str:
    return doc.category_workflow or ""


def _write_date_cell(ws: Worksheet, row: int, column: int, value: Optional[date]) -> None:
    if value is None:
        return
    cell = ws.cell(row=row, column=column, value=value)
    cell.number_format = DATE_NUMBER_FORMAT


def _apply_mdr_auto_filter(
    ws: Worksheet,
    last_data_row: int,
    *,
    last_col: str = FILTER_LAST_COL,
) -> None:
    """Filtri Excel su riga 13 e dati sotto (A13:last_col).

    B13:F13 e' un'intestazione unita: mostra un solo filtro sulla colonna B;
    nasconde i pulsante filtro su C-F (colId 2-5 rispetto ad A).
    Se last_col va oltre AI (es. colonne DBG_*), i filtri includono anche quelle.
    """
    if last_data_row < FILTER_HEADER_ROW:
        last_data_row = FILTER_HEADER_ROW
    ws.auto_filter.ref = (
        f"{FILTER_FIRST_COL}{FILTER_HEADER_ROW}:{last_col}{last_data_row}"
    )
    ws.auto_filter.filterColumn.clear()
    first_col_idx = ws[FILTER_FIRST_COL + "1"].column
    for col_idx in range(
        FILTER_MERGED_DESC_FIRST_COL + 1,
        FILTER_MERGED_DESC_LAST_COL + 1,
    ):
        col_id = col_idx - first_col_idx
        ws.auto_filter.filterColumn.append(
            FilterColumn(colId=col_id, hiddenButton=True)
        )


def _schedule_debug_last_col() -> str:
    return get_column_letter(COL_DBG_FIRST + len(SCHEDULE_DEBUG_HEADERS) - 1)


def _write_schedule_debug_header(
    ws: Worksheet,
    *,
    project_start: Optional[date],
) -> None:
    from openpyxl.styles import Alignment, Font

    last_dbg_col = COL_DBG_FIRST + len(SCHEDULE_DEBUG_HEADERS) - 1
    if project_start is not None:
        meta = (
            f"DBG — Project start: {project_start.isoformat()} "
            f"(PLANNED FIRST ISSUE col. AC when no predecessor shift)"
        )
        ws.merge_cells(
            start_row=SCHEDULE_DEBUG_META_ROW,
            start_column=COL_DBG_FIRST,
            end_row=SCHEDULE_DEBUG_META_ROW,
            end_column=last_dbg_col,
        )
        meta_cell = ws.cell(row=SCHEDULE_DEBUG_META_ROW, column=COL_DBG_FIRST, value=meta)
        meta_cell.font = Font(bold=True, italic=True)
        meta_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_font = Font(bold=True)
    for offset, header in enumerate(SCHEDULE_DEBUG_HEADERS):
        cell = ws.cell(row=FILTER_HEADER_ROW, column=COL_DBG_FIRST + offset, value=header)
        cell.font = header_font


def _write_schedule_debug_row(ws: Worksheet, row: int, doc: MdrLineItem) -> None:
    values: list = [
        doc.raci_title_key,
        doc.schedule_debug_pred_finishes,
        doc.schedule_debug_driving_pred,
        doc.planned_start,
        doc.duration_days if doc.duration_days is not None else "",
        doc.planned_finish,
        doc.schedule_debug_missing_preds,
        doc.schedule_debug_flags,
    ]
    for offset, value in enumerate(values):
        col = COL_DBG_FIRST + offset
        if value == "" or value is None:
            continue
        if offset in (_DBG_COL_PLANNED_START, _DBG_COL_PLANNED_FINISH):
            if isinstance(value, date):
                _write_date_cell(ws, row, col, value)
            continue
        ws.cell(row=row, column=col, value=safe_excel_value(value))


def write_mdr_excel(
    template_path: Path,
    output_path: Path,
    documents: List[MdrLineItem],
    project_code: Optional[str] = None,
    discipline_short_codes: Optional[Dict[str, str]] = None,
    *,
    schedule_debug_columns: bool = False,
    project_start: Optional[date] = None,
) -> Path:
    """
    Create a new xlsx from scratch: header copied from template, data rows written below.
    The template file is never copied as a whole, so printerSettings / print areas are absent.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tpl_wb = load_workbook(template_path, data_only=False)
    tpl_ws = tpl_wb[SHEET_NAME]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _copy_template_header(tpl_ws, ws)
    tpl_wb.close()

    if schedule_debug_columns:
        _write_schedule_debug_header(ws, project_start=project_start)

    proj = (project_code or "").strip() or "0000"
    pair_progress: Dict[tuple[str, str], int] = {}

    for idx, doc in enumerate(documents):
        row = DATA_START_ROW + idx
        disc_value = _line_discipline(doc)
        if discipline_short_codes:
            disc_value = discipline_short_codes.get(disc_value, disc_value)
        type_value = safe_excel_value(_line_type(doc))
        pair_key = (str(disc_value or ""), str(type_value or ""))
        pair_progress[pair_key] = pair_progress.get(pair_key, 0) + 1
        prog_value = f"{pair_progress[pair_key]:06d}"

        ws.cell(row=row, column=COL_ITEM, value=idx + 1)
        ws.cell(row=row, column=COL_DESCRIPTION, value=safe_excel_value(_line_title(doc)))
        ws.cell(row=row, column=COL_DISC, value=safe_excel_value(disc_value))
        ws.cell(row=row, column=COL_TYPE, value=type_value)
        ws.cell(row=row, column=COL_PROG, value=prog_value)
        ws.cell(row=row, column=COL_CATEGORY, value=safe_excel_value(_line_category(doc)))
        ws.cell(row=row, column=COL_WBS, value=safe_excel_value(_line_wbs(doc)))
        ws.cell(row=row, column=COL_WORKFLOW, value=safe_excel_value(_line_workflow(doc)))
        ws.cell(row=row, column=COL_FORMULA, value=_reneco_code_formula(row, proj))
        if doc.manhours is not None and doc.manhours >= 0:
            ws.cell(row=row, column=COL_MANHOURS, value=doc.manhours)
        _write_date_cell(ws, row, COL_PLANNED_FIRST_ISSUE, doc.planned_start)
        if schedule_debug_columns:
            _write_schedule_debug_row(ws, row, doc)

    last_data_row = DATA_START_ROW + len(documents) - 1 if documents else FILTER_HEADER_ROW
    filter_last = (
        _schedule_debug_last_col() if schedule_debug_columns else FILTER_LAST_COL
    )
    _apply_mdr_auto_filter(ws, last_data_row, last_col=filter_last)

    wb.properties.title = proj
    wb.save(output_path)
    wb.close()
    return output_path