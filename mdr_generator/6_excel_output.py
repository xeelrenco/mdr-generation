"""Step 6: Build Master Document Register Excel (clean workbook, no template copy)."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn
from openpyxl.worksheet.worksheet import Worksheet

from .models import MdrLineItem, SelectedDocument
from .utils import safe_excel_value

LineItemLike = Union[MdrLineItem, SelectedDocument]

DATA_START_ROW = 14
HEADER_LAST_ROW = 13
COL_ITEM = 1
COL_DESCRIPTION = 2  # B — unica colonna dati per intestazione unita B13:F13
COL_FORMULA = 7
COL_DISC = 9
COL_TYPE = 10
COL_PROG = 11  # K — progressive per coppia I+J
COL_CATEGORY = 15
COL_WBS = 21  # U
COL_MANHOURS = 24  # X — ore/uomo = giorni timeline × 8
COL_WORKFLOW = 27  # AA
COL_PLANNED_FIRST_ISSUE = 29  # AC — PLANNED FIRST ISSUE (AD non compilata per ora)
FILTER_HEADER_ROW = 13
FILTER_FIRST_COL = "A"
FILTER_LAST_COL = "AI"  # col 35 — KBS INTERNAL CODE
# Template: B13:F13 merged — filtro descrizione solo su colonna B (offset 1 da A).
FILTER_MERGED_DESC_FIRST_COL = COL_DESCRIPTION
FILTER_MERGED_DESC_LAST_COL = 6  # F
SHEET_NAME = "MDR"
DATE_NUMBER_FORMAT = "mm-dd-yy"


def _copy_cell_style(src, dst) -> None:
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def _copy_template_header(source: Worksheet, target: Worksheet, max_row: int = HEADER_LAST_ROW) -> None:
    """Copy rows 1..max_row (values, styles, merges, dimensions, images) — no print metadata."""
    max_col = max(source.max_column, COL_WORKFLOW)

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        dim = source.column_dimensions.get(letter)
        if dim is not None and dim.width is not None:
            target.column_dimensions[letter].width = dim.width

    for row_idx in range(1, max_row + 1):
        src_rd = source.row_dimensions.get(row_idx)
        if src_rd is not None and src_rd.height is not None:
            target.row_dimensions[row_idx].height = src_rd.height

    for merge in source.merged_cells.ranges:
        if merge.min_row <= max_row and merge.max_row <= max_row:
            target.merge_cells(str(merge))

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            src = source.cell(row=row_idx, column=col_idx)
            dst = target.cell(row=row_idx, column=col_idx, value=src.value)
            _copy_cell_style(src, dst)

    for image in getattr(source, "_images", ()):
        target.add_image(copy(image), image.anchor)


def _reneco_code_formula(row: int, project_prefix: str) -> str:
    """H (originator) is left empty for manual entry; K is the I+J pair progressive."""
    prefix = (project_prefix or "0000").replace('"', '""')
    return f'="{prefix}-"&H{row}&"-"&I{row}&J{row}&"-"&K{row}'


def _line_title(doc: LineItemLike) -> str:
    if isinstance(doc, MdrLineItem):
        return doc.mdr_document_title
    return doc.title


def _line_discipline(doc: LineItemLike) -> str:
    return doc.discipline_code


def _line_type(doc: LineItemLike) -> str:
    return doc.type_code or ""


def _line_category(doc: LineItemLike) -> str:
    return doc.category_code or ""


def _line_wbs(doc: LineItemLike) -> str:
    return doc.discipline_wbs or ""


def _line_workflow(doc: LineItemLike) -> str:
    return doc.category_workflow or ""


def _write_date_cell(ws: Worksheet, row: int, column: int, value: Optional[date]) -> None:
    if value is None:
        return
    cell = ws.cell(row=row, column=column, value=value)
    cell.number_format = DATE_NUMBER_FORMAT


def _apply_mdr_auto_filter(ws: Worksheet, last_data_row: int) -> None:
    """Filtri Excel su riga 13 e dati sotto (A13:AI).

    B13:F13 e' un'intestazione unita: mostra un solo filtro sulla colonna B;
    nasconde i pulsante filtro su C-F (colId 2-5 rispetto ad A).
    """
    if last_data_row < FILTER_HEADER_ROW:
        last_data_row = FILTER_HEADER_ROW
    ws.auto_filter.ref = (
        f"{FILTER_FIRST_COL}{FILTER_HEADER_ROW}:{FILTER_LAST_COL}{last_data_row}"
    )
    ws.auto_filter.filterColumn.clear()
    first_col_idx = ws[FILTER_FIRST_COL + "1"].column
    for col_idx in range(
        FILTER_MERGED_DESC_FIRST_COL + 1,
        FILTER_MERGED_DESC_LAST_COL + 1,
    ):
        col_id = col_idx - first_col_idx
        ws.auto_filter.filterColumn.append(
            FilterColumn(colId=col_id, hiddenButton=True)
        )


def write_mdr_excel(
    template_path: Path,
    output_path: Path,
    documents: List[LineItemLike],
    project_code: Optional[str] = None,
    discipline_short_codes: Optional[Dict[str, str]] = None,
) -> Path:
    """
    Create a new xlsx from scratch: header copied from template, data rows written below.
    The template file is never copied as a whole, so printerSettings / print areas are absent.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tpl_wb = load_workbook(template_path, data_only=False)
    tpl_ws = tpl_wb[SHEET_NAME]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _copy_template_header(tpl_ws, ws)
    tpl_wb.close()

    proj = (project_code or "").strip() or "0000"
    pair_progress: Dict[tuple[str, str], int] = {}

    for idx, doc in enumerate(documents):
        row = DATA_START_ROW + idx
        disc_value = _line_discipline(doc)
        if discipline_short_codes:
            disc_value = discipline_short_codes.get(disc_value, disc_value)
        type_value = safe_excel_value(_line_type(doc))
        pair_key = (str(disc_value or ""), str(type_value or ""))
        pair_progress[pair_key] = pair_progress.get(pair_key, 0) + 1
        prog_value = f"{pair_progress[pair_key]:06d}"

        ws.cell(row=row, column=COL_ITEM, value=idx + 1)
        ws.cell(row=row, column=COL_DESCRIPTION, value=safe_excel_value(_line_title(doc)))
        ws.cell(row=row, column=COL_DISC, value=safe_excel_value(disc_value))
        ws.cell(row=row, column=COL_TYPE, value=type_value)
        ws.cell(row=row, column=COL_PROG, value=prog_value)
        ws.cell(row=row, column=COL_CATEGORY, value=safe_excel_value(_line_category(doc)))
        ws.cell(row=row, column=COL_WBS, value=safe_excel_value(_line_wbs(doc)))
        ws.cell(row=row, column=COL_WORKFLOW, value=safe_excel_value(_line_workflow(doc)))
        ws.cell(row=row, column=COL_FORMULA, value=_reneco_code_formula(row, proj))
        if isinstance(doc, MdrLineItem):
            if doc.manhours is not None and doc.manhours >= 0:
                ws.cell(row=row, column=COL_MANHOURS, value=doc.manhours)
            _write_date_cell(ws, row, COL_PLANNED_FIRST_ISSUE, doc.planned_start)

    last_data_row = DATA_START_ROW + len(documents) - 1 if documents else FILTER_HEADER_ROW
    _apply_mdr_auto_filter(ws, last_data_row)

    wb.properties.title = proj
    wb.save(output_path)
    wb.close()
    return output_path
