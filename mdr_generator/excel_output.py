"""Step 6: Fill Master Document Register Excel template."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from .models import SelectedDocument
from .utils import safe_excel_value, sanitize_excel_for_open

DATA_START_ROW = 14
COL_ITEM = 1
COL_DESCRIPTION = 2
COL_DISC = 9
COL_TYPE = 10
COL_PROG = 11  # PROG.N0 — used by RENCO CODE formula
COL_FORMULA = 7
SHEET_NAME = "MDR"
PROJECT_CODE_CELL = "H14"  # project segment in RENCO CODE formula


def _clear_openpyxl_print_metadata(wb, sheet_name: str) -> None:
    ws = wb[sheet_name]
    ws.print_area = ""
    try:
        ws.print_title_rows = None
    except ValueError:
        pass
    try:
        ws.print_title_cols = None
    except ValueError:
        pass

    for name in list(wb.defined_names):
        key = name.lower()
        if "print" in key or key.startswith("_xlnm.print"):
            del wb.defined_names[name]


def _project_code_for_formula(wb, ws) -> str:
    """Value for H14 (project code segment in RENCO CODE)."""
    cell = PROJECT_CODE_CELL
    val = ws[cell].value
    if val is not None and str(val).strip():
        return str(val).strip()
    return wb.properties.title or "0000"


def _reneco_code_formula(row: int) -> str:
    """Build RENCO CODE formula; project code in $H$14; PROG.N0 defaults to 01 if empty."""
    return (
        f'="8360-"&$H$14&"-"&I{row}&J{row}&"-"&IF(K{row}="","01",K{row})'
    )


def write_mdr_excel(
    template_path: Path,
    output_path: Path,
    documents: List[SelectedDocument],
    project_code: Optional[str] = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    # Strip printer metadata from template copy before openpyxl touches it
    sanitize_excel_for_open(output_path)

    wb = load_workbook(output_path)
    ws = wb[SHEET_NAME]
    _clear_openpyxl_print_metadata(wb, SHEET_NAME)

    proj = project_code or _project_code_for_formula(wb, ws)
    ws[PROJECT_CODE_CELL] = proj

    for idx, doc in enumerate(documents):
        row = DATA_START_ROW + idx
        ws.cell(row=row, column=COL_ITEM, value=idx + 1)
        ws.cell(
            row=row,
            column=COL_DESCRIPTION,
            value=safe_excel_value(doc.title),
        )
        ws.cell(row=row, column=COL_DISC, value=safe_excel_value(doc.discipline_code))
        ws.cell(row=row, column=COL_TYPE, value=safe_excel_value(doc.type_code))
        ws.cell(row=row, column=COL_PROG, value=f"{idx + 1:02d}")
        ws.cell(row=row, column=COL_FORMULA, value=_reneco_code_formula(row))

    # Clear leftover template rows below data
    last_data_row = DATA_START_ROW + len(documents)
    for row in range(last_data_row, ws.max_row + 1):
        item_val = ws.cell(row=row, column=COL_ITEM).value
        if isinstance(item_val, (int, float)) and item_val > len(documents):
            for col in (COL_DESCRIPTION, COL_DISC, COL_TYPE, COL_PROG, COL_FORMULA):
                ws.cell(row=row, column=col, value=None)

    wb.save(output_path)
    sanitize_excel_for_open(output_path)
    return output_path
