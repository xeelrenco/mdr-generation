"""Step 7: Report qualità generazione MDR (leggibile, in italiano)."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    MdrLineItem,
    NormalizedSignal,
    PipelineSummary,
    RencoComparison,
    RawScopeSignal,
    SelectedDocument,
    UncertainMapping,
)
from .llm_usage import (
    LlmUsageSummary,
    format_cost_usd,
    format_token_count,
    provider_billing_label,
)
from .utils import format_elapsed_seconds, safe_excel_value

GeneratedDoc = Union[MdrLineItem, SelectedDocument]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D6DCE4")
SECTION_FONT = Font(bold=True, color="1F3864")
NOTE_FONT = Font(italic=True, color="595959")

EXCLUSION_REASON_IT = {
    "discipline_not_in_raci_vocabulary": "Disciplina non presente nel catalogo RACI",
    "chapter_not_in_raci_vocabulary": "Capitolo non presente nel catalogo RACI",
    "chapter_required_missing": "Capitolo mancante nel segnale LLM",
    "pair_not_in_catalog": "Coppia disciplina+capitolo assente dal catalogo documenti",
    "chapter_no_documents_in_catalog": "Capitolo noto ma senza documenti in catalogo",
    "source_pages_missing": "Pagine PDF mancanti nel segnale LLM",
    "source_pages_outside_chunk": "Pagine citate fuori dal chunk di estrazione",
}

CATEGORY_IT = {
    "overlap": "Presente in entrambi",
    "solo_generato": "Solo MDR generato (scope attuale)",
    "solo_renco_raci": "Solo MDR Renco (RACI) — gap di scope",
}


_CONSENSUS_RULE_IT = {
    "pass2_strong_direct": "Evidenza strong pass 2",
    "judges_agree": "Giudici concordi",
    "arbiter_decided": "Arbitro",
    "arbiter_no_verdict": "Arbitro senza verdetto",
    "fail_open_incomplete": "Fail-open (risposta incompleta)",
}


def _write_table(
    ws,
    start_row: int,
    headers: List[str],
    rows: List[List[Any]],
    col_widths: Optional[List[int]] = None,
) -> int:
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, start=start_row + 1):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=safe_excel_value(val))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    width_list = col_widths or [18] * len(headers)
    for ci, width in enumerate(width_list, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
    return start_row + len(rows) + 2


def _write_section_title(ws, row: int, title: str, merge_cols: int = 6) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    if merge_cols > 1:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=merge_cols
        )
    return row + 1


def _reason_it(reason: str) -> str:
    base = EXCLUSION_REASON_IT.get(reason.split(";")[0].strip(), reason)
    if ";" in reason:
        return f"{base} — {reason.split(';', 1)[1].strip()}"
    return base


def _build_gap_analysis_rows(
    renco: RencoComparison,
    normalized: List[NormalizedSignal],
    selected: List[GeneratedDoc],
) -> tuple[List[List[Any]], List[List[Any]]]:
    """Righe dettaglio gap + riepilogo per disciplina."""
    scope_pairs = {(n.discipline_code, n.chapter_name or "") for n in normalized}
    output_pairs = {
        (document.discipline_code, document.chapter_name or "")
        for document in selected
    }
    pair_renco_docs: dict[tuple[str, str], int] = {}
    for p in renco.scope_pairs:
        if p.present_in_renco_raci:
            pair_renco_docs[(p.discipline_code, p.chapter_name)] = (
                p.renco_documents_in_pair
            )

    detail: List[List[Any]] = []
    gap_pairs: set[tuple[str, str]] = set()
    for r in renco.detail_rows:
        if r.category != "solo_renco_raci":
            continue
        pair = (r.discipline_code, r.chapter_name or "")
        gap_pairs.add(pair)
        scope_state = (
            "Pair con output"
            if pair in output_pairs
            else "Pair estratta, 0 doc"
            if pair in scope_pairs
            else "Fuori scope"
        )
        detail.append(
            [
                r.discipline_code,
                r.chapter_name,
                r.raci_title,
                scope_state,
                pair_renco_docs.get(pair, "—"),
            ]
        )
    detail.sort(key=lambda x: (x[0], x[1], x[2]))

    pairs_with_output = sum(1 for pair in gap_pairs if pair in output_pairs)
    pairs_scope_no_output = sum(
        1 for pair in gap_pairs if pair in scope_pairs and pair not in output_pairs
    )
    summary = [
        ["Titoli RACI nel gap", len(detail), "Solo MDR Renco (MATCH)"],
        ["Coppie disc+cap distinte nel gap", len(gap_pairs), ""],
        [
            "Coppie gap con documenti in output",
            pairs_with_output,
            "Almeno un documento finale nella coppia",
        ],
        [
            "Coppie gap estratte ma con 0 documenti",
            pairs_scope_no_output,
            "Tutti i candidati rimossi da 2d/3e",
        ],
        [
            "Coppie gap ancora fuori scope",
            len(gap_pairs) - pairs_with_output - pairs_scope_no_output,
            "Candidati per second pass mirato sul PDF",
        ],
    ]
    by_disc = Counter(r[0] for r in detail)
    for disc, count in sorted(by_disc.items()):
        summary.append([f"  — {disc}", count, "Titoli nel gap per disciplina"])

    return summary, detail


def write_qa_report(
    output_path: Path,
    raw_signals: List[RawScopeSignal],
    normalized: List[NormalizedSignal],
    uncertain: List[UncertainMapping],
    selected: List[GeneratedDoc],
    summary: PipelineSummary,
    renco: Optional[RencoComparison] = None,
    llm_usage: Optional[LlmUsageSummary] = None,
    exclusion_audit: Optional[dict] = None,
    basis_gate_audit: Optional[dict] = None,
    consensus_audit: Optional[dict] = None,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # --- Foglio 1: Panoramica ---
    ws = wb.create_sheet("Panoramica", 0)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52

    row = 1
    ws.cell(row=row, column=1, value="Report qualità — MDR Generator").font = Font(
        bold=True, size=14
    )
    row += 2

    row = _write_section_title(ws, row, "Cosa contiene questo file")
    notes = [
        (
            "Panoramica",
            "",
            "Numeri chiave della run e confronto con MDR Renco (solo titoli RACI reconciliati).",
        ),
        (
            "Scope",
            "",
            "Coppie disciplina+capitolo estratte dal PDF e quanti documenti RACI generano.",
        ),
        (
            "Esclusi",
            "",
            "Segnali scope scartati in Step 2 (coppia non valida o fuori catalogo).",
        ),
        (
            "Esclusioni_SoW",
            "",
            "Ambiti committente/fuori scope (Step 2d), pair/documenti rimossi e "
            "documenti senza base nello SoW (Step 3e).",
        ),
        (
            "MDR_generato",
            "",
            "Elenco finale inserito nel Master Document Register.",
        ),
        (
            "Confronto_Renco",
            "",
            "Diff su titoli RACI: overlap, solo generato, solo Renco. I documenti NO_MATCH restano fuori.",
        ),
        (
            "Gap_analisi",
            "",
            "Titoli Renco non generati: verifica se la coppia scope è presente in questa run.",
        ),
    ]
    row = _write_table(
        ws, row, ["Foglio", "Valore", "Descrizione"], notes, [22, 12, 52]
    )

    row = _write_section_title(ws, row, "Pipeline — conteggi")
    pipeline_rows = [
        ["Progetto", summary.project_name, ""],
        ["PDF Scope", "; ".join(summary.scope_pdfs), ""],
        [
            "Provider LLM Scope",
            summary.scope_llm_provider or "—",
            "Pass 1 — estrazione scope",
        ],
        [
            "Modello LLM Scope",
            summary.scope_llm_model or "—",
            "Pass 1 — risolto da config/CLI",
        ],
        [
            "Pass 2 consenso catalogo",
            "Sì" if summary.scope_pass2_enabled else "No",
            "Verifica indipendente dell'intero catalogo + tie-break sui disaccordi",
        ],
        [
            "Provider / modello pass 2",
            (
                f"{summary.scope_pass2_provider} / {summary.scope_pass2_model}"
                if summary.scope_pass2_enabled
                else "—"
            ),
            "Config SCOPE_PASS2_*",
        ],
        [
            "Coppie verificate pass 2",
            summary.scope_pass2_pairs_verified if summary.scope_pass2_enabled else "—",
            "Intero catalogo RACI, in batch deterministici",
        ],
        [
            "Coppie finali dopo consenso",
            summary.scope_pass2_pairs_final if summary.scope_pass2_enabled else "—",
            "Accordo pass 1/pass 2, evidenza strong, giudici o arbitro",
        ],
        [
            "Giudici sulle coppie contese",
            (
                summary.scope_pass2_judges or "—"
                if summary.scope_pass2_enabled
                else "—"
            ),
            "Voto cieco sul PDF completo: se concordano la decisione è loro",
        ],
        [
            "Coppie contese",
            summary.scope_pass2_disagreements if summary.scope_pass2_enabled else "—",
            "Disaccordo pass 1/pass 2 o risposte pass 2 incomplete",
        ],
        [
            "Arbitro sui disaccordi tra giudici",
            (
                f"{summary.scope_pass2_arbiter_model or '—'} — "
                f"{summary.scope_pass2_judges_disagree} coppie, "
                f"{summary.scope_pass2_arbiter_present} ammesse"
                if summary.scope_pass2_enabled
                else "—"
            ),
            "Legge PDF completo e argomenti dei due giudici, poi decide",
        ],
        [
            "Ammesse da evidenza strong",
            (
                summary.scope_pass2_strong_direct
                if summary.scope_pass2_enabled
                else "—"
            ),
            "Non trovate dal pass 1 ma con evidenza strong: entrano senza giudici",
        ],
        [
            "Coppie senza verdetto dell'arbitro",
            (
                summary.scope_pass2_arbiter_no_verdict
                if summary.scope_pass2_enabled
                else "—"
            ),
            "Arbitro non ha risposto: restano solo se trovate dal pass 1",
        ],
        [
            "Fail-open per risposta incompleta",
            summary.scope_pass2_fallbacks if summary.scope_pass2_enabled else "—",
            "Pair mantenute perché il tie-break non ha dato una decisione completa",
        ],
        [
            "Stabilità vs run precedente (Jaccard)",
            (
                f"{summary.scope_stability_jaccard:.4f}"
                if summary.scope_stability_jaccard is not None
                else "—"
            ),
            (
                f"{summary.scope_stability_previous_run}; "
                f"+{summary.scope_stability_pairs_added} / "
                f"-{summary.scope_stability_pairs_removed} pair; "
                f"delta candidati="
                f"{summary.scope_stability_candidate_delta:+d}"
                if summary.scope_stability_previous_run
                and summary.scope_stability_candidate_delta is not None
                else (
                    f"{summary.scope_stability_previous_run}; "
                    f"+{summary.scope_stability_pairs_added} / "
                    f"-{summary.scope_stability_pairs_removed} pair"
                    if summary.scope_stability_previous_run
                    else "Prima run archiviata con il nuovo audit"
                )
            ),
        ],
        ["1. Segnali LLM (raw)", summary.raw_signal_count, "Estratti dal PDF"],
        [
            "2. Coppie scope valide",
            summary.normalized_signal_count,
            "Disciplina+capitolo presenti in catalogo (dopo Step 2d)",
        ],
        [
            "2d. Ambiti esclusi SoW",
            summary.scope_exclusions_active,
            "Committente / fuori scope con evidence",
        ],
        [
            "   — coppie rimosse",
            summary.scope_pairs_dropped,
            "Coppie disciplina|chapter eliminate dallo scope",
        ],
        [
            "   — documenti rimossi",
            summary.scope_docs_dropped,
            "TitleKey candidati esclusi per package SoW",
        ],
        [
            "   — documenti segnalati, non applicati",
            summary.scope_docs_flagged,
            (
                "Circuit breaker 2d scattato"
                if summary.scope_exclusion_guard_triggered
                else "—"
            ),
        ],
        [
            "3e. Doc senza base nello SoW",
            summary.sow_basis_docs_dropped,
            "Temi non previsti dal progetto (gate generalista)",
        ],
        [
            "   — 3e segnalati, non applicati",
            summary.sow_basis_docs_flagged,
            (
                "Circuit breaker 3e scattato"
                if summary.sow_basis_guard_triggered
                else "—"
            ),
        ],
        [
            "3. Candidati prima delle esclusioni",
            summary.candidates_before_exclusions,
            "Catalogo dalle coppie scope prima di 2d/3e",
        ],
        [
            "3. Candidati dopo 2d",
            summary.candidates_after_2d,
            "Dopo i quattro livelli di esclusione",
        ],
        ["3. Documenti RACI candidati", summary.candidate_count, "Da coppie scope"],
        [
            "3b. Decisioni document scope",
            summary.document_scope_decisions,
            "Pass LLM per coppia (in scope / istanze)",
        ],
        [
            "3d. Title enrichment (SoW)",
            "Sì" if summary.title_enrichment_enabled else "No",
            "Titoli SoW-specifici + split righe v2 (sow_elements)",
        ],
        [
            "   — doc con titolo SoW",
            summary.title_enrichment_docs_with_sow if summary.title_enrichment_enabled else "—",
            "Documenti con almeno un sow_specific_title",
        ],
        [
            "   — righe extra da split",
            summary.title_enrichment_extra_rows if summary.title_enrichment_enabled else "—",
            "Righe MDR oltre il conteggio post-3b",
        ],
        ["4. Righe MDR finali", summary.mdr_line_items or summary.selected_count, "Output Excel MDR"],
        [
            "   — durata timeline popolata",
            summary.duration_populated_count,
            "Giorni calendario (Finish−Start) da v_TimelineTaskToMdrLinks_Dates; usata per Step 5 schedule",
        ],
        [
            "   — MANHOURS popolati",
            summary.manhours_populated_count,
            "Colonna X: round(giorni timeline × 8); vuoto se manca durata timeline",
        ],
        [
            "   — schedule attivo",
            "Sì" if summary.schedule_enabled else "No",
            "Ordine per predecessor RACI",
        ],
        [
            "   — righe con PLANNED FIRST ISSUE",
            summary.schedule_dated_rows,
            "Colonna AC nel file MDR.xlsx",
        ],
        [
            "   — con storico MATCH",
            summary.with_history_count,
            "Già visti in reconciliation",
        ],
        [
            "   — senza storico",
            summary.without_history_count,
            "In scope ma senza MATCH storico",
        ],
        ["Segnali scope esclusi", summary.uncertain_mapping_count, "Vedi foglio Esclusi"],
        [
            "Tempo esecuzione totale",
            format_elapsed_seconds(summary.elapsed_seconds),
            f"{summary.elapsed_seconds:.1f} s — durata end-to-end pipeline",
        ],
        [
            "Stima costi LLM (totale)",
            format_cost_usd(summary.llm_estimated_cost_usd),
            (
                f"{format_token_count(summary.llm_total_input_tokens)} input + "
                f"{format_token_count(summary.llm_total_output_tokens)} output, "
                f"{summary.llm_total_calls} chiamate — stima USD"
            ),
        ],
    ]
    if llm_usage and llm_usage.provider_cost_usd:
        for provider, amount in sorted(
            llm_usage.provider_cost_usd.items(),
            key=lambda item: (-item[1], item[0]),
        ):
            pipeline_rows.append(
                [
                    f"   — {provider_billing_label(provider)}",
                    format_cost_usd(amount),
                    "Fatturazione separata per provider",
                ]
            )
    if llm_usage and llm_usage.lines:
        for line in llm_usage.lines:
            pipeline_rows.append(
                [
                    f"   — LLM {line.stage}",
                    format_cost_usd(line.cost_usd),
                    (
                        f"{line.model} ({line.call_type}): {line.calls} chiamate, "
                        f"{format_token_count(line.input_tokens)} in / "
                        f"{format_token_count(line.output_tokens)} out"
                    ),
                ]
            )
    if llm_usage and llm_usage.pricing_note:
        pipeline_rows.append(
            [
                "   — nota tariffe LLM",
                "",
                llm_usage.pricing_note,
            ]
        )
    row = _write_table(
        ws, row, ["Metrica", "Valore", "Nota"], pipeline_rows, [32, 14, 44]
    )

    if renco:
        row = _write_section_title(ws, row, "Confronto MDR Renco (solo RACI MATCH)")
        ws.cell(row=row, column=1, value="Fonte riferimento").font = NOTE_FONT
        ws.cell(row=row, column=2, value=renco.source_path)
        row += 1
        renco_rows = [
            [
                "Righe titolo nello storico progetto",
                renco.renco_rows_total,
                "v_MdrPreviousRecords_Normalized_All",
            ],
            [
                "Righe reconciliate MATCH → RACI",
                renco.renco_reconciled_match,
                "Usate nel confronto",
            ],
            [
                "Righe NO_MATCH (non RACI)",
                renco.renco_reconciled_no_match,
                "Escluse dal diff — normali in MDR progetto",
            ],
            [
                "Titoli RACI distinti nel Renco",
                renco.renco_raci_titles_distinct,
                "Dopo dedup su catalogo",
            ],
            ["Titoli RACI nel MDR generato", renco.generated_titles, ""],
            ["Overlap (in entrambi)", renco.overlap_count, "Allineamento OK"],
            [
                "Solo MDR generato",
                renco.only_generated_count,
                "In scope ora, non nel Renco mappato",
            ],
            [
                "Solo MDR Renco (RACI)",
                renco.only_renco_raci_count,
                "Gap di scope — candidati per ampliare estrazione PDF",
            ],
        ]
        _write_table(ws, row, ["Metrica", "Valore", "Nota"], renco_rows, [32, 14, 44])

    # --- Foglio 2: Scope ---
    ws = wb.create_sheet("Scope")
    pair_doc_counts = {}
    for s in selected:
        key = (s.discipline_code, s.chapter_name)
        pair_doc_counts[key] = pair_doc_counts.get(key, 0) + 1

    renco_pair_flags = {}
    if renco:
        for p in renco.scope_pairs:
            renco_pair_flags[(p.discipline_code, p.chapter_name)] = (
                p.present_in_renco_raci,
                p.renco_documents_in_pair,
            )

    scope_rows = []
    for n in normalized:
        pair = (n.discipline_code, n.chapter_name or "")
        in_renco, renco_n = renco_pair_flags.get(pair, (None, None))
        scope_rows.append(
            [
                n.discipline_code,
                n.chapter_name or "",
                n.scope_section,
                pair_doc_counts.get(pair, 0),
                "Sì" if in_renco else ("No" if in_renco is False else "—"),
                renco_n if renco_n is not None else "—",
                (
                    "; ".join(
                        f"{source_pdf}: {','.join(str(page) for page in pages)}"
                        for source_pdf, pages in sorted(
                            n.source_pages_by_pdf.items()
                        )
                    )
                    if n.source_pages_by_pdf
                    else ",".join(str(p) for p in n.source_pages)
                ),
                (n.notes or "")[:300],
            ]
        )
    _write_table(
        ws,
        1,
        [
            "Disciplina",
            "Capitolo",
            "Sezione SoW",
            "Doc in MDR",
            "Coppia in Renco RACI",
            "Doc Renco (coppia)",
            "Pagine PDF",
            "Evidenza (estratto)",
        ],
        scope_rows,
        [10, 28, 24, 10, 16, 14, 12, 40],
    )

    # --- Foglio 3: Esclusi ---
    ws = wb.create_sheet("Esclusi")
    if uncertain:
        excl_rows = [
            [
                u.scope_section,
                u.raw_discipline,
                u.raw_chapter,
                _reason_it(u.reason),
                u.source_pdf,
            ]
            for u in uncertain
        ]
    else:
        excl_rows = [["—", "—", "—", "Nessun segnale escluso in questa run", "—"]]
    _write_table(
        ws,
        1,
        ["Sezione SoW", "Disciplina LLM", "Capitolo LLM", "Motivo", "PDF"],
        excl_rows,
        [24, 12, 28, 36, 28],
    )

    # --- Foglio 2b: Scope_consenso (Step 2c) ---
    ws = wb.create_sheet("Scope_consenso")
    row = _write_section_title(
        ws, 1, "Coppie non decise dall'accordo pass 1 / pass 2 (Step 2c)"
    )
    decisions = [
        item
        for item in ((consensus_audit or {}).get("pair_decisions") or [])
        if item.get("resolution") != "agreement"
    ]
    consensus_rows = [
        [
            item.get("discipline_code", ""),
            item.get("chapter_name", ""),
            _CONSENSUS_RULE_IT.get(
                item.get("resolution", ""), item.get("resolution", "")
            ),
            "Presente" if item.get("final_decision") == "present" else "Esclusa",
            item.get("pass1_vote", ""),
            item.get("pass2_vote", ""),
            item.get("pass2_support_chunks", ""),
            "Sì" if item.get("pass2_has_strong") else "No",
            item.get("judge1_vote", ""),
            item.get("judge2_vote", ""),
            item.get("arbiter_vote", ""),
            (item.get("arbiter_reason") or item.get("judge1_reason") or "")[:300],
        ]
        for item in decisions
    ] or [["—"] * 12]
    _write_table(
        ws,
        row,
        [
            "Disciplina",
            "Capitolo",
            "Chi ha deciso",
            "Esito",
            "Pass 1",
            "Pass 2",
            "Conferme pass 2",
            "Evidenza strong",
            "Giudice 1",
            "Giudice 2",
            "Arbitro",
            "Motivazione",
        ],
        consensus_rows,
        [10, 26, 22, 10, 10, 10, 12, 12, 12, 12, 12, 46],
    )

    # --- Foglio 3b: Esclusioni_SoW (Step 2d) ---
    ws = wb.create_sheet("Esclusioni_SoW")
    excl_audit = exclusion_audit or {}
    row = _write_section_title(ws, 1, "Ambiti esclusi dallo SoW (Step 2d)")
    excl_items = excl_audit.get("exclusions") or []
    if excl_items:
        package_rows = [
            [
                e.get("label", e.get("package", "")),
                e.get("exclude_level", ""),
                e.get("application_status", ""),
                e.get("responsibility", ""),
                "Sì" if e.get("explicit_assuntore") else "No",
                e.get("exclusion_type", ""),
                ",".join(e.get("discipline_codes") or e.get("suggested_discipline_codes") or []),
                "; ".join(e.get("chapter_names") or []),
                "; ".join(
                    f"{p.get('discipline_code')}|{p.get('chapter_name')}"
                    for p in (e.get("pairs") or [])
                    if isinstance(p, dict)
                ),
                "Sì" if e.get("should_exclude") else "No",
                e.get("confidence", ""),
                "; ".join(e.get("parse_warnings") or []),
                (e.get("evidence_quote") or "")[:300],
                "; ".join(e.get("source_pdfs") or [e.get("source_pdf", "")]),
            ]
            for e in excl_items
        ]
    else:
        package_rows = [
            [
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "Nessuna esclusione trovata",
                "—",
            ]
        ]
    row = _write_table(
        ws,
        row,
        [
            "Label",
            "Livello",
            "Stato applicazione",
            "Responsabilità",
            "Assuntore esplicito",
            "Tipo",
            "Discipline",
            "Chapter names",
            "Pair RACI",
            "Attiva",
            "Conf.",
            "Warning",
            "Evidence",
            "PDF",
        ],
        package_rows,
        [22, 12, 20, 14, 14, 18, 14, 28, 36, 8, 8, 38, 40, 22],
    )
    row += 1
    row = _write_section_title(ws, row, "Coppie rimosse")
    dropped_pairs = excl_audit.get("dropped_pairs") or []
    if dropped_pairs:
        pair_rows = [
            [
                d.get("discipline_code", ""),
                d.get("chapter_name", ""),
                d.get("exclude_level", ""),
                d.get("label", d.get("package", "")),
                d.get("reason", ""),
                (d.get("evidence_quote") or "")[:250],
            ]
            for d in dropped_pairs
        ]
    else:
        pair_rows = [["—", "—", "—", "—", "Nessuna coppia rimossa", "—"]]
    row = _write_table(
        ws,
        row,
        ["Disciplina", "Capitolo", "Livello", "Label", "Motivo", "Evidence"],
        pair_rows,
        [12, 36, 12, 22, 22, 40],
    )
    row += 1
    row = _write_section_title(ws, row, "Documenti RACI rimossi")
    dropped_docs = excl_audit.get("dropped_documents") or []
    if dropped_docs:
        doc_rows = [
            [
                d.get("discipline_code", ""),
                d.get("chapter_name", ""),
                d.get("title", ""),
                d.get("title_key", ""),
                d.get("exclude_level", ""),
                d.get("label", d.get("package", "")),
                d.get("reason", ""),
            ]
            for d in dropped_docs
        ]
    else:
        doc_rows = [["—", "—", "—", "—", "—", "—", "Nessun documento rimosso"]]
    row = _write_table(
        ws,
        row,
        ["Disciplina", "Capitolo", "Titolo", "TitleKey", "Livello", "Label", "Motivo"],
        doc_rows,
        [12, 28, 40, 28, 12, 18, 22],
    )

    if excl_audit.get("drop_guard_triggered"):
        row += 1
        row = _write_section_title(
            ws, row, "2d segnalati ma non applicati (circuit breaker)"
        )
        flagged_docs = excl_audit.get("flagged_documents") or []
        flagged_rows = [
            [
                d.get("discipline_code", ""),
                d.get("chapter_name", ""),
                d.get("title", ""),
                d.get("title_key", ""),
                d.get("exclude_level", ""),
                d.get("label", ""),
                d.get("llm_reason", d.get("reason", "")),
            ]
            for d in flagged_docs
        ] or [["—", "—", "—", "—", "—", "—", "Nessun dettaglio"]]
        row = _write_table(
            ws,
            row,
            ["Disciplina", "Capitolo", "Titolo", "TitleKey", "Livello", "Label", "Motivo"],
            flagged_rows,
            [12, 28, 40, 28, 12, 18, 30],
        )

    doc_llm_rows = excl_audit.get("document_llm_audit") or []
    invalid_doc_rows = [
        [
            item.get("title_key", ""),
            item.get("outcome", ""),
            item.get("exclusion_label", ""),
            item.get("source_pdf", ""),
            str(item.get("raw", ""))[:300],
        ]
        for item in doc_llm_rows
        if item.get("outcome") != "excluded"
    ]
    if invalid_doc_rows:
        row += 1
        row = _write_section_title(ws, row, "Audit mapping document-level")
        row = _write_table(
            ws,
            row,
            ["TitleKey", "Esito", "Esclusione", "PDF", "Raw"],
            invalid_doc_rows,
            [30, 20, 28, 24, 50],
        )

    row += 1
    row = _write_section_title(
        ws, row, "Documenti senza base nello SoW (Step 3e)"
    )
    gate_audit = basis_gate_audit or {}
    gate_docs = gate_audit.get("dropped_documents") or []
    if gate_audit.get("discarded_excessive_drop"):
        gate_rows = [
            [
                d.get("discipline_code", ""),
                d.get("chapter_name", ""),
                d.get("title", ""),
                d.get("title_key", ""),
                "NON APPLICATO — " + (d.get("reason", "") or "circuit breaker"),
            ]
            for d in (gate_audit.get("flagged_documents") or [])
        ] or [
            [
                "—",
                "—",
                "—",
                "—",
                f"Risultato scartato: {gate_audit.get('documents_flagged', 0)} documenti "
                f"segnalati su {gate_audit.get('candidates_before', 0)}",
            ]
        ]
    elif gate_docs:
        gate_rows = [
            [
                d.get("discipline_code", ""),
                d.get("chapter_name", ""),
                d.get("title", ""),
                d.get("title_key", ""),
                d.get("reason", ""),
            ]
            for d in gate_docs
        ]
    else:
        gate_rows = [["—", "—", "—", "—", "Nessun documento rimosso"]]
    _write_table(
        ws,
        row,
        ["Disciplina", "Capitolo", "Titolo", "TitleKey", "Motivo"],
        gate_rows,
        [12, 28, 40, 28, 46],
    )

    # --- Foglio 4: MDR_generato ---
    ws = wb.create_sheet("MDR_generato")
    mdr_rows = []
    for s in selected:
        if isinstance(s, MdrLineItem):
            mdr_rows.append(
                [
                    s.discipline_code,
                    s.chapter_name,
                    s.mdr_document_title,
                    s.raci_title,
                    s.sow_specific_title,
                    s.sow_title_confidence,
                    s.raci_title_key,
                    s.instance_count,
                    "Sì" if s.scalable else "No",
                    s.duration_days if s.duration_days is not None else "",
                    s.manhours if s.manhours is not None else "",
                    s.planned_start.isoformat() if s.planned_start else "",
                    s.historical_count,
                    "Sì" if s.bucket == "with_history" else "No",
                    s.category_code,
                    s.type_code,
                ]
            )
        else:
            mdr_rows.append(
                [
                    s.discipline_code,
                    s.chapter_name,
                    s.title,
                    s.title,
                    s.title_key,
                    1,
                    "—",
                    "",
                    "",
                    s.historical_count,
                    "Sì" if s.bucket == "with_history" else "No",
                    s.category_code,
                    s.type_code,
                ]
            )
    _write_table(
        ws,
        1,
        [
            "Disciplina",
            "Capitolo",
            "Titolo display (col B)",
            "Titolo RACI",
            "Titolo SoW",
            "Conf. SoW",
            "TitleKey",
            "Istanze",
            "Scalable",
            "Giorni (timeline)",
            "MANHOURS",
            "Planned First Issue",
            "Occorrenze storico",
            "Storico MATCH",
            "Category",
            "Type",
        ],
        mdr_rows,
        [10, 26, 44, 36, 36, 10, 14, 8, 10, 12, 12, 14, 14, 12, 10, 10],
    )

    # --- Foglio 5: Confronto_Renco ---
    ws = wb.create_sheet("Confronto_Renco")
    row = 1
    if not renco:
        ws.cell(
            row=1,
            column=1,
            value="Confronto non disponibile — connessione MotherDuck o storico progetto assente.",
        )
    else:
        ws.cell(row=row, column=1, value="Legenda categorie").font = SECTION_FONT
        row += 1
        for cat, label in CATEGORY_IT.items():
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=label)
            row += 1
        row += 1

        def _detail_rows(category: str) -> List[List[Any]]:
            return [
                [
                    CATEGORY_IT.get(r.category, r.category),
                    r.discipline_code,
                    r.chapter_name,
                    r.raci_title,
                    r.title_key,
                    r.historical_count if category != "solo_renco_raci" else "",
                ]
                for r in renco.detail_rows
                if r.category == category
            ]

        for category in ("overlap", "solo_generato", "solo_renco_raci"):
            items = _detail_rows(category)
            row = _write_section_title(ws, row, CATEGORY_IT[category])
            if items:
                row = _write_table(
                    ws,
                    row,
                    [
                        "Categoria",
                        "Disciplina",
                        "Capitolo",
                        "Titolo RACI",
                        "TitleKey",
                        "Storico gen.",
                    ],
                    items,
                    [28, 10, 26, 44, 14, 12],
                )
            else:
                ws.cell(row=row, column=1, value="(nessuna riga)")
                row += 2

    # --- Foglio 6: Gap_analisi ---
    ws = wb.create_sheet("Gap_analisi")
    row = 1
    if not renco:
        ws.cell(
            row=1,
            column=1,
            value="Gap non disponibile — confronto Renco assente.",
        )
    else:
        gap_summary, gap_detail = _build_gap_analysis_rows(
            renco, normalized, selected
        )
        row = _write_section_title(ws, row, "Riepilogo gap vs MDR Renco")
        row = _write_table(
            ws, row, ["Metrica", "Valore", "Nota"], gap_summary, [36, 10, 44]
        )
        row = _write_section_title(ws, row, "Titoli Renco non generati (solo_renco_raci)")
        if gap_detail:
            row = _write_table(
                ws,
                row,
                [
                    "Disciplina",
                    "Capitolo",
                    "Titolo RACI",
                    "Stato scope run",
                    "Doc Renco (coppia)",
                ],
                gap_detail,
                [10, 28, 44, 22, 16],
            )
        else:
            ws.cell(row=row, column=1, value="(nessun gap — overlap completo)")
            row += 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
