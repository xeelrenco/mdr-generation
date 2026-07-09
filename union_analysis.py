#!/usr/bin/env python3
"""
Analisi post-hoc: unione MDR di più run vs storico Renco (per TitleKey RACI).

Script standalone — NON fa parte della pipeline run_mdr_generator.py.
Eseguilo manualmente dopo una o più run, quando vuoi confrontare benchmark LLM.

Legge i fogli MDR_generato e Confronto_Renco dai generation_report.xlsx
(prodotti dallo Step 7) e calcola overlap/gap per singoli run e unioni a coppie.
Il confronto usa TitleKey RACI, come renco_compare.py — indipendente dall'arricchimento
titoli (Step 3d) che modifica solo il titolo display.

Usage:
  python union_analysis.py report_a.xlsx report_b.xlsx
  python union_analysis.py output/*_generation_report.xlsx
  python union_analysis.py --output-dir output --all-pairs
  python union_analysis.py --pair run1+run2=7350_20260521_172938+7350_20260521_153457
"""

from __future__ import annotations

import argparse
import itertools
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Set, Tuple

import openpyxl

from mdr_generator.config import PROJECT_DIR

CATEGORY_OVERLAP = "Presente in entrambi"
CATEGORY_RENCO_GAP = "Solo MDR Renco (RACI) — gap di scope"

MDR_GENERATO_COL_TITLE_KEY = 7
CONFRONTO_RENCO_COL_TITLE_KEY = 5


def _norm_key(raw: object) -> str | None:
    if raw is None:
        return None
    key = str(raw).strip().lower()
    if not key or key == "—":
        return None
    return key


def _load_generated_title_keys(report_path: Path) -> Set[str]:
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    ws = wb["MDR_generato"]
    keys: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        key = _norm_key(ws.cell(r, MDR_GENERATO_COL_TITLE_KEY).value)
        if key:
            keys.add(key)
    wb.close()
    return keys


def _section_title_keys(report_path: Path, section_label: str) -> Set[str]:
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    ws = wb["Confronto_Renco"]
    keys: Set[str] = set()
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() != section_label:
            continue
        key = _norm_key(ws.cell(r, CONFRONTO_RENCO_COL_TITLE_KEY).value)
        if key:
            keys.add(key)
    wb.close()
    return keys


def _load_renco_reference_keys(report_path: Path) -> Set[str]:
    """Tutti i TitleKey RACI MATCH dello storico: overlap + gap di scope."""
    overlap = _section_title_keys(report_path, CATEGORY_OVERLAP)
    gap = _section_title_keys(report_path, CATEGORY_RENCO_GAP)
    return overlap | gap


def _stats(generated: Set[str], renco: Set[str]) -> Tuple[int, int, int, int]:
    overlap = generated & renco
    return len(generated), len(overlap), len(renco) - len(overlap), len(generated - renco)


def _report_label(path: Path) -> str:
    name = path.name
    if name.endswith("_generation_report.xlsx"):
        return name[: -len("_generation_report.xlsx")]
    return path.stem


def _resolve_reports(
    output_dir: Path,
    explicit: List[str] | None,
) -> Dict[str, Path]:
    if explicit:
        out: Dict[str, Path] = {}
        for p in explicit:
            path = Path(p)
            if not path.is_absolute():
                path = output_dir / path
            if not path.exists():
                raise FileNotFoundError(path)
            label = _report_label(path)
            if label in out:
                label = f"{label}@{path.parent.name}"
            out[label] = path
        return out

    out: Dict[str, Path] = {}
    paths = sorted(
        output_dir.glob("*_generation_report.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for path in paths:
        out[_report_label(path)] = path
    return out


def _print_single(label: str, gen: Set[str], renco: Set[str]) -> Tuple[int, int, int, int]:
    mdr, ov, gap, extra = _stats(gen, renco)
    pct = round(100 * ov / len(renco), 1) if renco else 0.0
    print(
        f"{label:36} keys={mdr:4}  overlap={ov:3} ({pct}%)  "
        f"gap={gap:3}  extra={extra}"
    )
    return mdr, ov, gap, extra


def _print_union(
    label: str,
    gen_a: Set[str],
    gen_b: Set[str],
    renco: Set[str],
    ov_a: int,
    ov_b: int,
) -> None:
    union = gen_a | gen_b
    mdr, ov, gap, extra = _stats(union, renco)
    pct = round(100 * ov / len(renco), 1) if renco else 0.0
    delta = ov - max(ov_a, ov_b)
    shared = len(gen_a & gen_b)
    print(
        f"{label:36} keys={mdr:4}  overlap={ov:3} ({pct}%)  gap={gap:3}  "
        f"delta=+{delta}  shared={shared}"
    )


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Confronto unioni MDR run vs storico Renco (per TitleKey RACI)"
    )
    parser.add_argument(
        "reports",
        nargs="*",
        help="Path ai generation_report.xlsx (default: tutti in --output-dir)",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_DIR / "output"),
        help="Directory output (default: output/)",
    )
    parser.add_argument(
        "--pair",
        nargs="*",
        metavar="LABEL=A+B",
        help="Coppie da valutare, es. pro+gpt=run_a+run_b",
    )
    parser.add_argument(
        "--all-pairs",
        action="store_true",
        help="Calcola tutte le coppie tra i report indicati",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    output_dir = Path(args.output_dir)
    try:
        reports = _resolve_reports(output_dir, args.reports or None)
    except FileNotFoundError as e:
        print(f"ERRORE: report non trovato: {e}", file=sys.stderr)
        return 1

    if len(reports) < 2:
        print(
            "ERRORE: servono almeno 2 generation_report.xlsx "
            "(passali come argomenti o mettili in --output-dir).",
            file=sys.stderr,
        )
        return 1

    ref_path = next(iter(reports.values()))
    renco = _load_renco_reference_keys(ref_path)
    generated = {label: _load_generated_title_keys(p) for label, p in reports.items()}

    print(
        f"Riferimento Renco: {len(renco)} TitleKey RACI MATCH "
        f"(da {ref_path.name})"
    )
    print("Confronto per TitleKey — righe split/arricchimento titoli non alterano il match.\n")
    print("=== SINGOLI ===")
    overlaps: Dict[str, int] = {}
    for label in sorted(reports.keys()):
        _, ov, _, _ = _print_single(label, generated[label], renco)
        overlaps[label] = ov

    print("\n=== UNIONI ===")
    union_done = False
    if args.all_pairs:
        union_done = True
        for a, b in itertools.combinations(sorted(reports.keys()), 2):
            _print_union(
                f"{a} + {b}",
                generated[a],
                generated[b],
                renco,
                overlaps[a],
                overlaps[b],
            )
    elif args.pair:
        union_done = True
        for spec in args.pair:
            if "=" in spec:
                label, rhs = spec.split("=", 1)
            else:
                label, rhs = spec, spec
            parts = [p.strip() for p in rhs.split("+")]
            if len(parts) != 2:
                print(f"ERRORE: pair invalido {spec!r}", file=sys.stderr)
                return 1
            a, b = parts
            if a not in generated or b not in generated:
                print(f"ERRORE: run sconosciuta in {spec!r}", file=sys.stderr)
                return 1
            _print_union(
                label,
                generated[a],
                generated[b],
                renco,
                overlaps[a],
                overlaps[b],
            )
    elif len(reports) == 2:
        union_done = True
        a, b = sorted(reports.keys())
        _print_union(
            f"{a} + {b}",
            generated[a],
            generated[b],
            renco,
            overlaps[a],
            overlaps[b],
        )

    if not union_done:
        print("(nessuna unione calcolata: usa --all-pairs o --pair LABEL=A+B)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
